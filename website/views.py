from flask import Blueprint, render_template, request, flash, jsonify, redirect, url_for, make_response, Response
from flask_login import login_required, current_user
from .models import WebToonIP, User, WebNovel, Books
from . import db
import json, datetime, pandas, random, shutil
from werkzeug.utils import secure_filename
from sqlalchemy import or_, desc
from os import walk
from pandas import ExcelWriter
import io, math
import pandas as pd
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

BACKUP_PATH = 'website/static/backup/'

Developer = '고재혁'

cj_colors = ['#ff5a00', '#decaa5', '#00bcc8',
             '#ff8347', '#ff898d', '#f2b6ac',
             '#88e6f2', '#009c91', '#cd004d',
             '#790029', '#ffde47', '#7b28f6',
             '#2fe530', '#cce3b9', '#11d8ad',
             '#00780c', '#37481e', '#013c40',
             '#17142f', '#3b1500', '#8d5700'
            ]
columns = {'num': 'NO',
        'title': '제목',
        'genre': '주장르',
        'keyword': '키워드',
        'ref_path': '추천경로',
        'copyright': '저작권',
        'writer': '원작자',
        'copyright_status': '저작권현황',
        'condition': '조건',
        'date_pub': '연재일/출판일',
        'monitored': '모니터링 여부',
        'date_monitored': '모니터링 시점',
        'suggested': '현업제안 여부',
        'date_suggested': '현업제안 시점',
        'story': '줄거리',
        'feedback': '검토의견'}

current_ips = []

prohibit_words = ['None', '']

db_ip_models = [WebToonIP, WebNovel, Books]

db_dict = {'WebToonIP': WebToonIP, 'WebNovel' : WebNovel, 'Books':Books}

views = Blueprint('views', __name__)

@views.route('/', methods=['GET', 'POST'])
@login_required
def home():
    ip_dict = {}
    for ip_model in db_ip_models:
        ips, total_num, ago = get_db_info(ip_model)
        ip_dict[ip_model] = [ips, total_num, ago]
    return render_template("home.html", user=current_user, ips = ip_dict)

def get_db_info(ip_model):
    ips = ip_model.query.order_by(desc(ip_model.date)).all()
    total_num = len(ips)
    current_time = datetime.datetime.now()
    date_time_str = ips[0].date
    date_time_obj = datetime.datetime.strptime(date_time_str, "%Y-%m-%d %H:%M:%S")
    ts = current_time - date_time_obj
    mins = int(ts.seconds/60)
    if ts.days:
        ago = str(ts.days) + ' day(s)'
    elif mins > 60:
        hours = int(mins/60)
        ago = str(hours) + ' hour(s)'
    elif mins > 0:
        ago = str(mins) + ' min(s)'
    else:
        ago = str(ts.seconds) + ' seconds'
    return ips, total_num, ago

@views.route('/<ip_type>', methods=['GET', 'POST'])
@login_required
def show_ip(ip_type):
    global current_ips
    current_db = db_dict[ip_type]
    all_genre = db.session.query(current_db.genre).distinct()
    all_copyright_status = db.session.query(current_db.copyright_status).distinct()
    all_keyword = db.session.query(current_db.keyword).distinct()
    this_year = datetime.datetime.now().year
    ips = current_db.query.order_by(desc(current_db.id)).all()
    if request.method == 'POST':
        old_ip_type = current_ips[0].__class__.__name__
        if not current_ips:
            current_ips = ips
        else:
            if ip_type != old_ip_type:
                current_ips = ips
        return export_ips(current_ips)
    total_num = len(ips)
    return render_template("webtoon_card.html", user=current_user, ips = ips, total_num = total_num, all_genre = all_genre, searched='False', all_copyright_status = all_copyright_status, all_keyword = all_keyword, this_year = this_year)

def export_ips(ips):
    wb_columns = ['NO','제목', '주장르','키워드','추천경로','저작권','원작자','저작권현황','조건','연재일/출판일','모니터링 여부','모니터링 시점','현업제안 여부', '현업제안 시점','줄거리', '검토의견']
    ip_type = ips[0].__class__.__name__
    current_time = datetime.datetime.now()
    current_time = str(current_time.timestamp()).split('.')[0]
    
    wb = Workbook(write_only = True)
    ws = wb.create_sheet(ip_type)
    
    if ip_type == 'WebToonIP':
        ws.append(wb_columns)
        for ip in ips:
            row=[ip.num,
                 ip.title,
                 ip.genre,
                 ip.keyword,
                 ip.ref_path,
                 ip.copyright,
                 ip.writer,
                 ip.copyright_status,
                 ip.condition,
                 ip.date_pub,
                 ip.monitored,
                 ip.date_monitor,
                 ip.suggested,
                 ip.date_suggested,
                 ip.story,
                 ip.feedback
                 ]
            ws.append(row)
    elif ip_type == 'WebNovel':
        wb_columns.insert(-2,'현업제안 내용')
        ws.append(wb_columns)
        for ip in ips:
            row=[ip.num,
                 ip.title,
                 ip.genre,
                 ip.keyword,
                 ip.ref_path,
                 ip.copyright,
                 ip.writer,
                 ip.copyright_status,
                 ip.condition,
                 ip.date_pub,
                 ip.monitored,
                 ip.date_monitor,
                 ip.suggested,
                 ip.date_suggested,
                 ip.suggested_feedback,
                 ip.story,
                 ip.feedback
                 ]
            ws.append(row)
    elif ip_type == 'Books':
        wb_columns.insert(2,'분류')
        wb_columns.insert(3,'국가')
        wb_columns.insert(-2,'현업제안 내용')
        ws.append(wb_columns)
        for ip in ips:
            row=[ip.num,
                 ip.title,
                 ip.category,
                 ip.nation,
                 ip.genre,
                 ip.keyword,
                 ip.ref_path,
                 ip.copyright,
                 ip.writer,
                 ip.copyright_status,
                 ip.condition,
                 ip.date_pub,
                 ip.monitored,
                 ip.date_monitor,
                 ip.suggested,
                 ip.date_suggested,
                 ip.suggested_feedback,
                 ip.story,
                 ip.feedback
                 ]
            ws.append(row)
    resp = Response(save_virtual_workbook(wb))
    # Set the response header to let the browser resolve to the file download behavior
    resp.headers['Content-Disposition'] = 'attachement; filename= ' + ip_type + '_' + current_time +'.xlsx'
    resp.headers['Content-Type'] = 'application/vnd.ms-excel; charset=utf-8'

    return resp

@views.route('/search/<ip_type>', methods=['GET', 'POST'])
@login_required
def search_ips(ip_type):
    global current_ips
    current_db = db_dict[ip_type]
    total_num = len(current_db.query.all())
    all_genre = db.session.query(current_db.genre).distinct()
    all_copyright_status = db.session.query(current_db.copyright_status).distinct()
    all_keyword = db.session.query(current_db.keyword).distinct()
    this_year = datetime.datetime.now().year
    if request.method == 'GET':
        search_title = "%{}%".format(request.args.get('searchtitle'))
        dm_year = request.args.get('search_dm_year')
        dm_month = request.args.get('search_dm_month')
        if dm_year != '' or dm_month != '':
            if dm_month != '':
                dm_month = '.' + dm_month  
            search_date_monitor = '%' + dm_year + dm_month + '%'
        else:
            search_date_monitor = Developer
        dp_year = request.args.get('search_dp_year')
        dp_month = request.args.get('search_dp_month')
        if dp_year != '' or dp_month != '':
            if dp_month != '':
                dp_month = '.' + dp_month
            search_date_pub = '%' + dp_year + dp_month + '%'
        else:
            search_date_pub = Developer
        search_genre = request.args.get('search_genre')
        search_copyright_status = request.args.get('search_copyright_status')
        search_keyword = "%{}%".format(request.args.get('search_keyword'))
        
        if search_title == '%%':
            search_title = Developer
        if search_genre == '':
            search_genre = Developer
        if search_copyright_status == '':
            search_copyright_status = Developer
        if search_keyword == '%%':
            search_keyword = Developer
        #print ('title:',search_title,'genre:', search_genre,'copyright:', search_copyright_status,'keyword:', search_keyword, '연재일:', search_date_pub, '모니터링 시점 :', search_date_monitor)
        results = current_db.query.order_by(desc(current_db.id)).filter(or_(current_db.title.like(search_title),
                                             current_db.date_monitor.like(search_date_monitor),
                                             current_db.date_pub.like(search_date_pub),
                                             current_db.genre.like(search_genre),
                                             current_db.copyright_status.like(search_copyright_status),
                                             current_db.keyword.like(search_keyword)
                                            )
                                        ).all()
        total_sum = len(results)
        if total_sum != 0:
            current_ips = results
            return render_template("webtoon_card.html", user=current_user, ips = current_ips, total_sum = total_sum, total_num = total_num, all_genre=all_genre, searched='True', all_copyright_status = all_copyright_status, all_keyword = all_keyword, this_year=this_year)
        else:
            flash('검색결과가 없습니다. 검색어를 확인해주세요!', category='error')
            current_ips = []
            return redirect('/'+ip_type)
    else:
        return redirect('/'+ip_type)
'''
@views.route('/WebNovel', methods=['GET', 'POST'])
@login_required
def webnovel():
    ips = WebNovel.query.order_by(desc(WebNovel.id)).all()
    all_genre = db.session.query(WebNovel.genre).distinct()
    all_copyright_status = db.session.query(WebNovel.copyright_status).distinct()
    all_keyword = db.session.query(WebNovel.keyword).distinct()
    this_year = datetime.datetime.now().year
    total_num = len(ips)
    return render_template("webtoon_card.html", user=current_user, ips = ips, total_num = total_num, all_genre = all_genre, searched='False', all_copyright_status = all_copyright_status, all_keyword = all_keyword, this_year = this_year)

@views.route('/Books', methods=['GET', 'POST'])
@login_required
def books():
    ips = Books.query.order_by(desc(Books.id)).all()
    all_genre = db.session.query(Books.genre).distinct()
    all_copyright_status = db.session.query(Books.copyright_status).distinct()
    all_keyword = db.session.query(Books.keyword).distinct()
    this_year = datetime.datetime.now().year
    total_num = len(ips)
    return render_template("webtoon_card.html", user=current_user, ips = ips, total_num = total_num, all_genre = all_genre, searched='False', all_copyright_status = all_copyright_status, all_keyword = all_keyword, this_year = this_year)
'''

@views.route('/components', methods=['GET', 'POST'])
@login_required
def components():
    users=User.query.all()
    ips_webtoon = WebToonIP.query.all()
    webtoon_copyright_status = db.session.query(WebToonIP.copyright_status).distinct()
    webtoon_genre = db.session.query(WebToonIP.genre).distinct()
    total_webtoon = len(ips_webtoon)
    return render_template("components.html", user=current_user, users=users, ips_webtoon = ips_webtoon, total_num = total_webtoon, webtoon_copyright_status=webtoon_copyright_status, webtoon_genre=webtoon_genre)
'''
@views.route("/export/<ip_type_search>", methods=['GET'])
@login_required
def export_search_ips(ip_type_search):
    ip_type = ip_type_search.split(':')[0]
    searched = ip_type_search.split(':')[-1]
    current_db = db_dict[ip_type]
    current_time = datetime.datetime.now()
    current_time = str(current_time.timestamp()).split('.')[0]
    all_ips = current_db.query.filter(current_db.id)
    # Instantiate byte type IO objects, used to store objects in memory, no need to generate temporary files on disk
    out = io.BytesIO()
    # Instantiate the writer object that outputs xlsx
    writer = ExcelWriter(out, engine='openpyxl')
    # Split the SQLAlchemy model query object into SQL statements and connection attributes to pandas read_sql method
    df = pd.read_sql(all_ips.statement, all_ips.session.bind)
    # Simple data slicing, select all rows, the range from the sixth column to the last column
    df = df.iloc[:, 1:-2]

    if ip_type == 'WebNovel':
        columns['suggested_feedback'] = '현업제안 내용'
    elif ip_type == 'Books':
        columns['category'] = '분류'
        columns['nation'] = '국가'
        columns['suggested_feedback'] = '현업제안 내용'
    # Rename the df column name
    df.rename(columns = columns, inplace=True)

    # Save df to excel in the memory writer variable, do not include the index line number in the conversion result
    df.to_excel(writer, index=False)
    # This step can't be missed, if you don't save it, there is nothing in the xls file downloaded by the browser
    writer.save()
    # Reset the pointer of the IO object to the beginning
    out.seek(0)
    # The IO object uses getvalue() to return the binary raw data, which is used to give the response data to be generated
    resp = make_response(out.getvalue())
    # Set the response header to let the browser resolve to the file download behavior
    resp.headers['Content-Disposition'] = 'attachement; filename=' + ip_type + '_' + current_time +'.xlsx'
    resp.headers['Content-Type'] = 'application/vnd.ms-excel; charset=utf-8'

    return resp
'''
'''
@views.route("/export/<int:results>", methods=['GET'])
@login_required
def export_searched_ips(results):
    ip_type = results[0].__class__.__name__
    current_time = datetime.datetime.now()
    current_time = str(current_time.timestamp()).split('.')[0]
    
    wb = Workbook(write_only = True)
    ws = wb.create_sheet(ip_type)
    ws.append(wb_column)
    
    resp = Response(save_virtual_workbook(wb))
    # Set the response header to let the browser resolve to the file download behavior
    resp.headers['Content-Disposition'] = 'attachement; filename= ' + ip_type + '_' + current_time +'.xlsx'
    resp.headers['Content-Type'] = 'application/vnd.ms-excel; charset=utf-8'

    return resp
'''

@views.route('/backup', methods=['GET', 'POST'])
@login_required
def backup():
    process = ''
    db_file = 'database.db'
    _, _, backup_list = next(walk(BACKUP_PATH))
    backup_list = list(reversed(sorted(backup_list)))
    if request.method == 'POST':
        '''copy current db to /backup/db_file'''
        process = request.form.get('button')
        if process == 'backup':
            ct=datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            db_file = 'database_' + ct
            shutil.copy('website/database.db', 'website/static/backup/%s.db' %db_file)
        elif process == 'restore':
            db_file = request.form.get('backupfiles')
    return render_template("backup.html", user=current_user, process = process, db_file = db_file, backup_list = backup_list)

def progress(status, remaining, total):
    print(f'Copied {total-remaining} of {total} pages...')

@views.route('/analytics', methods=['GET', 'POST'])
@login_required
def analytics():
    ips = WebToonIP.query.all()
    total_num = len(ips)
    all_genre = db.session.query(WebToonIP.genre).distinct()
    
    genre_list = get_list(all_genre)
    genre_list.append('genre')
    genre_dict = get_dict(genre_list,'WebToonIP')
    genre_graph_data = get_key_value_list(genre_dict)
    
    all_copyright_status = db.session.query(WebToonIP.copyright_status).distinct()
    copyright_status_list = get_list(all_copyright_status)
    copyright_status_list.append('copyright_status')
    copyright_status_dict = get_dict(copyright_status_list,'WebToonIP')
    copyright_status_graph_data = get_key_value_list(copyright_status_dict)
    
    all_date_monitor = db.session.query(WebToonIP.date_monitor).distinct()
    date_monitor_list = get_list(all_date_monitor)
    date_monitor_list.append('date_monitor')
    date_monitor_dict = get_dict(date_monitor_list,'WebToonIP')
    date_monitor_graph_data = get_key_value_list_date_monitor(date_monitor_dict)
    #ips_webnovel = WebNovel.query.order_by(desc(WebNovel.date)).all()
    ips_webnovel = WebNovel.query.all()
    total_num_webnovel = len(ips_webnovel)
    all_genre_webnovel = db.session.query(WebNovel.genre).distinct()
    
    genre_list_webnovel = get_list(all_genre_webnovel)
    genre_list_webnovel.append('genre')
    genre_dict_webnovel = get_dict(genre_list_webnovel,'WebNovel')
    genre_graph_data_webnovel = get_key_value_list(genre_dict_webnovel)

    all_copyright_status_webnovel = db.session.query(WebNovel.copyright_status).distinct()
    copyright_status_list_webnovel = get_list(all_copyright_status_webnovel)
    copyright_status_list_webnovel.append('copyright_status')
    copyright_status_dict_webnovel = get_dict(copyright_status_list_webnovel,'WebNovel')
    copyright_status_graph_data_webnovel = get_key_value_list(copyright_status_dict_webnovel)
    
    all_date_monitor_webnovel = db.session.query(WebNovel.date_monitor).distinct()
    date_monitor_list_webnovel = get_list(all_date_monitor_webnovel)
    date_monitor_list_webnovel.append('date_monitor')
    date_monitor_dict_webnovel = get_dict(date_monitor_list_webnovel,'WebNovel')
    date_monitor_graph_data_webnovel = get_key_value_list_date_monitor(date_monitor_dict_webnovel)
    
    ips_books = Books.query.all()
    total_num_books = len(ips_books)
    all_genre_books = db.session.query(Books.genre).distinct()
    
    genre_list_books = get_list(all_genre_books)
    genre_list_books.append('genre')
    genre_dict_books = get_dict(genre_list_books,'Books')
    genre_graph_data_books = get_key_value_list(genre_dict_books)

    all_copyright_status_books = db.session.query(Books.copyright_status).distinct()
    copyright_status_list_books = get_list(all_copyright_status_books)
    copyright_status_list_books.append('copyright_status')
    copyright_status_dict_books = get_dict(copyright_status_list_books,'Books')
    copyright_status_graph_data_books = get_key_value_list(copyright_status_dict_books)
    
    all_date_monitor_books = db.session.query(Books.date_monitor).distinct()
    date_monitor_list_books = get_list(all_date_monitor_books)
    date_monitor_list_books.append('date_monitor')
    date_monitor_dict_books = get_dict(date_monitor_list_books,'Books')
    date_monitor_graph_data_books = get_key_value_list_date_monitor(date_monitor_dict_books)
    
    return render_template("analytics.html", 
                           user=current_user,
                           ips = ips,
                           total = total_num,
                           all_genre = genre_dict,
                           all_copyright_status = copyright_status_dict,
                           all_date_monitor = date_monitor_dict,
                           genre_graph_data = genre_graph_data,
                           copyright_status_graph_data = copyright_status_graph_data,
                           date_monitor_graph_data = date_monitor_graph_data,
                           ips_webnovel = ips_webnovel,
                           total_webnovel = total_num_webnovel,
                           all_genre_webnovel = genre_dict_webnovel,
                           genre_graph_data_webnovel = genre_graph_data_webnovel,
                           all_copyright_status_webnovel = copyright_status_dict_webnovel,
                           copyright_status_graph_data_webnovel = copyright_status_graph_data_webnovel,
                           all_date_monitor_webnovel = date_monitor_dict_webnovel,
                           date_monitor_graph_data_webnovel = date_monitor_graph_data_webnovel,
                           ips_books = ips_books,
                           total_books = total_num_books,
                           all_genre_books = genre_dict_books,
                           genre_graph_data_books = genre_graph_data_books,
                           all_copyright_status_books = copyright_status_dict_books,
                           copyright_status_graph_data_books = copyright_status_graph_data_books,
                           all_date_monitor_books = date_monitor_dict_books,
                           date_monitor_graph_data_books = date_monitor_graph_data_books
                          )
def get_key_value_list_date_monitor(dict_):
    list_ = [['ITEM', 'COUNT',{ 'role' : 'style' }, { 'role' : 'annotation' }]]
    sorted_list = []
    temp_sorted = []
    for key in dict_.keys():
        if key.split('.')[-1].isdigit():
            sorted_list.append(key)
            #list_.append([key, dict_[key], cj_colors[int(random.random()*len(cj_colors))], dict_[key]])
    for i in sorted_list:
        if len(i.split('.')[-1]) == 2:
            temp=i.split('.')[0]+i.split('.')[-1]
        else:
            temp=i.split('.')[0]+ '0' + i.split('.')[-1]
        temp_sorted.append(int(temp))
    #sorted_list = sorted(temp_sorted)
    new_sorted_list=[]
    for i in sorted(temp_sorted):
        if str(i)[-2:][0] == '0':
            new_sorted_list.append(str(i)[:4]+'.'+str(i)[-1:])
        else:
            new_sorted_list.append(str(i)[:4]+'.'+str(i)[-2:])
    sorted_list = new_sorted_list
    for key in dict_.keys():
        if not key.split('.')[-1].isdigit():
            sorted_list.insert(0, key)
    for key in sorted_list:
        list_.append([str(key), dict_[str(key)], cj_colors[int(random.random()*len(cj_colors))], dict_[str(key)]])
    return list_

def get_key_value_list(dict_):
    list_ = [['ITEM', 'COUNT',{ 'role' : 'style' }, { 'role' : 'annotation' }]]
    res = sorted(dict_.items(), key=(lambda x:x[1]), reverse = True)
    
    for key,value in res:
        list_.append([key,value, cj_colors[int(random.random()*len(cj_colors))], value])
    return list_
    
def find_key(dict_, value):
    for k in dict_.keys():
        if dict_[k] == value:
            return k
    
def get_dict(item_list, IP):
    item_dict = {}
    if item_list[-1] == 'genre':
        item_list.pop()
        if IP == 'WebToonIP':
            for item in item_list:
                item_dict[item] = len(WebToonIP.query.filter(WebToonIP.genre.like('%' + item + '%')).all())
        elif IP == 'WebNovel':
            for item in item_list:
                item_dict[item] = len(WebNovel.query.filter(WebNovel.genre.like('%' + item + '%')).all())
        elif IP == 'Books':
            for item in item_list:
                item_dict[item] = len(Books.query.filter(Books.genre.like('%' + item + '%')).all())
    elif item_list[-1] == 'copyright_status':
        item_list.pop()
        if IP == 'WebToonIP':
            for item in item_list:
                item_dict[item] = len(WebToonIP.query.filter(WebToonIP.copyright_status.like('%' + item + '%')).all())
        elif IP == 'WebNovel':
            for item in item_list:
                item_dict[item] = len(WebNovel.query.filter(WebNovel.copyright_status.like('%' + item + '%')).all())
        elif IP == 'Books':
            for item in item_list:
                item_dict[item] = len(Books.query.filter(Books.copyright_status.like('%' + item + '%')).all())
    elif item_list[-1] == 'date_monitor':
        item_list.pop()
        if IP == 'WebToonIP':
            for item in item_list:
                if item != '.':
                    if item[-1] == '.':
                        item_dict[item[:-1]] = len(WebToonIP.query.filter(WebToonIP.date_monitor.like('%' + item[:-1] + '%')).all())
                    else:
                        item_dict[item] = len(WebToonIP.query.filter(WebToonIP.date_monitor.like('%' + item + '%')).all())
                        #print (item, len(WebToonIP.query.filter(WebToonIP.date_monitor.like('%' + item + '%')).all()))
        elif IP == 'WebNovel':
            for item in item_list:
                if item != '.':
                    if item[-1] == '.':
                        item_dict[item[:-1]] = len(WebNovel.query.filter(WebNovel.date_monitor.like('%' + item[:-1] + '%')).all())
                    else:
                        item_dict[item] = len(WebNovel.query.filter(WebNovel.date_monitor.like('%' + item + '%')).all())
        elif IP == 'Books':
            for item in item_list:
                if item != '.':
                    if item[-1] == '.':
                        item_dict[item[:-1]] = len(Books.query.filter(Books.date_monitor.like('%' + item[:-1] + '%')).all())
                    else:
                        item_dict[item] = len(Books.query.filter(Books.date_monitor.like('%' + item + '%')).all())
    return item_dict

def get_list(item_list):
    items_list = []
    for item in item_list:
        if item[0]:
            items_list.append(str(item[0]))
        else:
            items_list.append(str('None'))
    return items_list

@views.route('/import-db-webtoon', methods=['GET','POST'])
def import_db_webtoon():
    if request.method == 'POST':
        #print(request.files['file'])
        f = request.files['file']
        data_xls = pandas.read_excel(f, engine = 'openpyxl', sheet_name=0,
                                    index_col='NO')
        for i in range(len(data_xls.index)):
            ip_exist = False
            show = data_xls.values[i]
            t,w = show[0], show[5]
            if len(str(t)) < 1:
                print('제목이 너무 짧습니다! 확인해주세요!', category='error')
            elif len(str(w)) < 1:
                print('원작자가 너무 짧습니다! 확인해주세요!', category='error')
            if WebToonIP.query.filter_by(title=t).first():
                title_db = WebToonIP.query.filter_by(title=t).all()
                for ttl in title_db:
                    if ttl.writer == w:
                        ip_exist = True
                        print (ttl.title, 'by', w ,' 동일한 작가의 작품이 있습니다! 확인해주세요!')
            if not ip_exist:
                #try:
                n = int(data_xls.index[i].item())
                g = show[1].split()[0]
                kw = show[2]
                rf = show[3]
                c = show[4]
                cs = show[6]
                condi =show[7]
                dp = str(show[8])[:4] + '.' + str(show[9])[:2]
                if not math.isnan(show[11]) and str(int(show[11])).isdigit():
                    m='Y'
                    dm_year = str(show[11])[:4]
                    if not math.isnan(show[12]):
                        dm_mon = str(show[12])[:2]
                    else:
                        dm_mon='NA'
                    dm = dm_year + '.' + dm_mon
                    print (t,'is monitored')
                else:
                    m='N'
                    dm='NA'
                    print (t, 'is NOT monitored')
                if not math.isnan(show[14]) and str(int(show[14])).isdigit():
                    sg='Y'
                    dsg_year = str(show[14])[:4]
                    dsg_mon = str(show[15])[:2]
                    dsg = dsg_year + '.' + dsg_mon
                    print (t,'is suggested')
                else:
                    sg='N'
                    dsg='NA'
                    print (t, 'is NOT suggested')

                st = show[16]
                fb = data_xls.values[i][17]

                new_ip = WebToonIP(num=n,
                                   title = t,
                                   genre = g,
                                   keyword = kw,
                                   ref_path=rf,
                                   copyright = c,
                                   writer = w,
                                   copyright_status = cs,
                                   condition = condi,
                                   date_pub=dp,
                                   monitored = m,
                                   date_monitor = dm,
                                   suggested = sg,
                                   date_suggested = dsg,
                                   story=st,
                                   feedback = fb,
                                   user_id=current_user.id, 
                                   date = str(datetime.datetime.now()).split('.')[0])
                db.session.add(new_ip)
                db.session.commit()
                print (type(n),n,t, g, kw,rf,c,w,cs,condi,dp,m,dm,sg,dsg,st,fb)
        print ('WEBTOON DB updated !!')
    return '''
    <!doctype html>
    <title>Upload an excel file</title>
    <h1>웹툰 Excel file upload</h1>
    <form action="" method=post enctype=multipart/form-data>
    <p><input type=file name=file><input type=submit value=Upload>
    </form>
    '''

@views.route('/import-db-webnovel', methods=['GET','POST'])
def import_db_webnovel():
    if request.method == 'POST':
        #print(request.files['file'])
        f = request.files['file']
        data_xls = pandas.read_excel(f, engine = 'openpyxl', sheet_name=0,
                                    index_col='오')
        for i in range(len(data_xls.index)):
            ip_exist = False
            show = data_xls.values[i]
            t,w = show[0], show[5]
            if len(str(t)) < 1:
                ip_exist = True
                print('제목이 너무 짧습니다! 확인해주세요!', category='error')
            elif len(str(w)) < 1:
                ip_exist = True
                print('원작자가 너무 짧습니다! 확인해주세요!', category='error')
            elif WebNovel.query.filter_by(title=t).first():
                title_db = WebNovel.query.filter_by(title=t).all()
                for ttl in title_db:
                    if ttl.writer == w:
                        ip_exist = True
                        print (ttl.title, 'by', w ,' 동일한 작가의 작품이 있습니다! 확인해주세요!')
            if not ip_exist:
                #try:
                n = int(data_xls.index[i].item())
                if type(show[1]) != float:
                    g = show[1].split()[0]
                else:
                    g='NA'
                kw = show[2]
                rf = show[3]
                c = show[4]
                cs = show[6]
                condi =show[7]
                dp = str(show[8])[:4] + '.' + str(show[9])[:2]

                if str(show[11]).isdigit():
                    m='Y'
                    dm_year = str(show[11])[:4]
                    dm_mon = str(show[12])[:2]
                    dm = dm_year + '.' + dm_mon
                    print (t,'is monitored')
                else:
                    m='N'
                    dm='NA'
                    print (t,'is NOT monitored')

                if str(show[14]).isdigit():
                    sg='Y'
                    dsg_year = str(int(show[14]))[:4]
                    dsg_mon = str(show[15])[:2]
                    dsg = dsg_year + '.' + dsg_mon
                    print (t,'is suggested')
                else:
                    sg='N'
                    dsg='NA'
                    print (t,'is NOT suggested')

                sf = show[16]
                st = show[17]
                fb = data_xls.values[i][18]
                new_ip = WebNovel(num=n,
                                   title = t,
                                   genre = g,
                                   keyword = kw,
                                   ref_path=rf,
                                   copyright = c,
                                   writer = w,
                                   copyright_status = cs,
                                   condition = condi,
                                   date_pub=dp,
                                   monitored = m,
                                   date_monitor = dm,
                                   suggested = sg,
                                   date_suggested = dsg,
                                   suggested_feedback = sf,
                                   story=st,
                                   feedback = fb,
                                   user_id=current_user.id, 
                                   date = str(datetime.datetime.now()).split('.')[0])
                db.session.add(new_ip)
                db.session.commit()
                print (type(n),n,t,  g, kw,rf,c,w,cs,condi,dp,m,dm,sg,dsg,sf, st,fb)
        print ('WebNovel DB updated')
    return '''
    <!doctype html>
    <title>Upload an excel file</title>
    <h1>웹소설 Excel file upload</h1>
    <form action="" method=post enctype=multipart/form-data>
    <p><input type=file name=file><input type=submit value=Upload>
    </form>
    '''

@views.route('/import-db-books', methods=['GET','POST'])
def import_db_books():
    if request.method == 'POST':
        f = request.files['file']
        data_xls = pandas.read_excel(f, engine = 'openpyxl', sheet_name=0,
                                    index_col='NO')
        for i in range(len(data_xls.index)):
            mp=False
            sp = False
            ip_exist = False
            show = data_xls.values[i]
            t,w = show[0], show[7]
            if len(str(t)) < 1:
                ip_exist = True
                print('제목이 너무 짧습니다! 확인해주세요!', category='error')
            elif len(str(w)) < 1:
                ip_exist = True
                print('원작자가 너무 짧습니다! 확인해주세요!', category='error')
            elif Books.query.filter_by(title=t).first():
                title_db = Books.query.filter_by(title=t).all()
                for ttl in title_db:
                    if ttl.writer == w:
                        ip_exist = True
                        print (ttl.title, 'by', w ,' 동일한 작가의 작품이 있습니다! 확인해주세요!')
            if not ip_exist:
                n = int(data_xls.index[i].item())
                cate = show[1]
                nat = show[2]
                if isinstance(show[3],float):
                    g='NA'
                else:
                    g = show[3].split()[0]
                kw = show[4]
                rf = show[5]
                c = show[6]
                cs = show[8]
                condi =show[9]
                dp = str(show[10])[:4] + '.' + str(show[11])[:2]
                if isinstance(show[13],float):
                    if show[13].is_integer():
                        mp=True
                elif isinstance(show[13],int):
                    mp=True
                if mp:
                    m='Y'
                    dm_year = str(show[13])[:4]
                    dm_mon = str(show[14])[:2]
                    dm = dm_year + '.' + dm_mon
                    print (t,'is monitored')
                else:
                    m='N'
                    dm='NA'
                    print (t,'is NOT monitored')
                if isinstance(show[16],float):
                    if show[16].is_integer():
                        sp=True
                elif isinstance(show[16],int):
                    sp=True
                if sp:
                    sg='Y'
                    dsg_year = str(show[16])[:4]
                    dsg_mon = str(show[17])[:2]
                    dsg = dsg_year + '.' + dsg_mon
                    print (t,'is suggested')
                else:
                    sg='N'
                    dsg='NA'
                    print (t, 'is NOT Suggested')

                sf = show[18]
                st = show[19]
                fb = show[20]

                new_ip = Books(num=n,
                                   title = t,
                                   category = cate,
                                   nation = nat,
                                   genre = g,
                                   keyword = kw,
                                   ref_path=rf,
                                   copyright = c,
                                   writer = w,
                                   copyright_status = cs,
                                   condition = condi,
                                   date_pub=dp,
                                   monitored = m,
                                   date_monitor = dm,
                                   suggested = sg,
                                   date_suggested = dsg,
                                   suggested_feedback = sf,
                                   story=st,
                                   feedback = fb,
                                   user_id=current_user.id, 
                                   date = str(datetime.datetime.now()).split('.')[0])
                db.session.add(new_ip)
                db.session.commit()
                print (type(n),n,t, cate, nat, g, kw,rf,c,w,cs,condi,dp,m,dm,sg,dsg,sf, st,fb)
        print ('Books DB updated')
    return '''
    <!doctype html>
    <title>Upload an excel file</title>
    <h1>출판도서 Excel file upload</h1>
    <form action="" method=post enctype=multipart/form-data>
    <p><input type=file name=file><input type=submit value=Upload>
    </form>
    '''

@views.route('/detail-view/WebToonIP/<int:id>', methods=['GET','POST'])
@login_required
def deatil_view_webtoon(id):
    ip_to_update = WebToonIP.query.get_or_404(id)
    return render_template("detail_view.html", user=current_user, ip = ip_to_update)

@views.route('/detail-view/WebNovel/<int:id>', methods=['GET','POST'])
@login_required
def deatil_view_webnovel(id):
    ip_to_update = WebNovel.query.get_or_404(id)
    return render_template("detail_view.html", user=current_user, ip = ip_to_update)

@views.route('/detail-view/Books/<int:id>', methods=['GET','POST'])
@login_required
def deatil_view_books(id):
    ip_to_update = Books.query.get_or_404(id)
    return render_template("detail_view.html", user=current_user, ip = ip_to_update)

@views.route('/adding/WebToonIP', methods=['GET','POST'])
@login_required
def adding():
    this_year = datetime.datetime.now().year
    ips = WebToonIP.query.order_by(desc(WebToonIP.num)).all()
    max_num = get_maxnum(ips)
    all_genre = db.session.query(WebToonIP.genre).distinct()
    all_ref_path = db.session.query(WebToonIP.ref_path).distinct()
    all_copyright_status = db.session.query(WebToonIP.copyright_status).distinct()
    if request.method == 'POST':
        ip_exist = False
        n = request.form.get('num')
        t = request.form.get('title')
        g = request.form.get('genre')
        g_added = request.form.get('genre_added')
        if not g and g_added:
            g = g_added
        if g and g_added:
            flash('주장르를 중복으로 선택/작성 했습니다! 해당 IP에서 확인해주세요!', category='error')
        kw = request.form.get('keyword')
        rf = request.form.get('ref_path')
        rf_added = request.form.get('ref_path_added')
        if not rf and rf_added:
            rf = rf_added
        if rf and rf_added:
            flash('추천경로를 중복으로 선택/작성 했습니다! 해당 IP에서 확인해주세요!', category='error')
        c = request.form.get('copyright')
        w = request.form.get('writer')
        cs = request.form.get('copyright_status')
        cs_added = request.form.get('copyright_status_added')
        if not cs and cs_added:
            cs = cs_added
        if cs and cs_added:
            flash('저작권/판권현황를 중복으로 선택/작성 했습니다! 해당 IP에서 확인해주세요!', category='error')
        condi = request.form.get('condition')
        dp = str(request.form.get('dp_year')) + '.' + str(request.form.get('dp_month'))
        m = request.form.get('monitored')
        dm = str(request.form.get('dm_year')) + '.' + str(request.form.get('dm_month'))
        sg = request.form.get('suggested')
        dsg = str(request.form.get('dsg_year')) + '.' + str(request.form.get('dsg_month'))
        st = request.form.get('story')
        fb = request.form.get('feedback')

        if len(t) < 1:
            flash('제목이 너무 짧습니다! 확인해주세요!', category='error')
        elif len(w) < 1:
            flash('원작자가 너무 짧습니다! 확인해주세요!', category='error')
        if WebToonIP.query.filter_by(title=t).first():
            title_db = WebToonIP.query.filter_by(title=t).all()
            for ttl in title_db:
                if ttl.writer == w:
                    ip_exist = True
                    flash(f'{ttl.title} by {w} 동일한 작가의 작품이 있습니다! 확인해주세요!', category='error')
        if not ip_exist:
            new_ip = WebToonIP(num=n,
                               title = t,
                               genre = g,
                               keyword = kw,
                               ref_path=rf,
                               copyright = c,
                               writer = w,
                               copyright_status = cs,
                               condition = condi,
                               date_pub=dp,
                               monitored = m,
                               date_monitor = dm,
                               suggested = sg,
                               date_suggested = dsg,
                               story=st,
                               feedback = fb,
                               user_id=current_user.id, 
                               date = str(datetime.datetime.now()).split('.')[0])
            db.session.add(new_ip)
            db.session.commit()
            flash('New WebToonIP added!', category='success')
            return redirect('/WebToonIP')
    return render_template("adding_ip.html", user=current_user, ips=ips, max_num = max_num, all_genre = all_genre, all_ref_path = all_ref_path, all_copyright_status = all_copyright_status, this_year = this_year)

def get_maxnum(ips):
    max = 0
    for ip in ips:
        if int(ip.num) > max:
            max = int(ip.num)
    return max+1

@views.route('/adding/WebNovel', methods=['GET','POST'])
@login_required
def adding_webnovel():
    this_year = datetime.datetime.now().year
    ips = WebNovel.query.order_by(desc(WebNovel.num)).all()
    all_genre = db.session.query(WebNovel.genre).distinct()
    all_ref_path = db.session.query(WebNovel.ref_path).distinct()
    all_copyright_status = db.session.query(WebNovel.copyright_status).distinct()
    max_num = get_maxnum(ips)
    if request.method == 'POST':
        ip_exist = False
        n = request.form.get('num')
        t = request.form.get('title')
        g = request.form.get('genre')
        g_added = request.form.get('genre_added')
        if not g and g_added:
            g = g_added
        elif g and g_added:
            flash('주장르를 중복으로 선택/작성 했습니다! 해당 IP에서 확인해주세요!', category='error')
        kw = request.form.get('keyword')
        rf = request.form.get('ref_path')
        rf_added = request.form.get('ref_path_added')
        if not rf and rf_added:
            rf = rf_added
        elif rf and rf_added:
            flash('추천경로를 중복으로 선택/작성 했습니다! 해당 IP에서 확인해주세요!', category='error')
        c = request.form.get('copyright')
        w = request.form.get('writer')
        cs = request.form.get('copyright_status')
        cs_added = request.form.get('copyright_status_added')
        if not cs and cs_added:
            cs = cs_added
        elif cs and cs_added:
            flash('저작권/판권현황를 중복으로 선택/작성 했습니다! 해당 IP에서 확인해주세요!', category='error')
        condi = request.form.get('condition')
        dp = str(request.form.get('dp_year')) + '.' + str(request.form.get('dp_month'))
        m = request.form.get('monitored')
        dm = str(request.form.get('dm_year')) + '.' + str(request.form.get('dm_month'))
        sg = request.form.get('suggested')
        dsg = str(request.form.get('dsg_year')) + '.' + str(request.form.get('dsg_month'))
        sf = request.form.get('suggested_feedback')
        st = request.form.get('story')
        fb = request.form.get('feedback')

        if len(t) < 1:
            flash('제목이 너무 짧습니다! 확인해주세요!', category='error')
        elif len(w) < 1:
            flash('원작자가 너무 짧습니다! 확인해주세요!', category='error')
        if WebNovel.query.filter_by(title=t).first():
            title_db = WebNovel.query.filter_by(title=t).all()
            for ttl in title_db:
                if ttl.writer == w:
                    ip_exist = True
                    flash(f'{ttl.title} by {w} 동일한 작가의 작품이 있습니다! 확인해주세요!', category='error')
        if not ip_exist:
            new_ip = WebNovel(num=n,
                               title = t,
                               genre = g,
                               keyword = kw,
                               ref_path=rf,
                               copyright = c,
                               writer = w,
                               copyright_status = cs,
                               condition = condi,
                               date_pub=dp,
                               monitored = m,
                               date_monitor = dm,
                               suggested = sg,
                               date_suggested = dsg,
                               suggested_feedback = sf,
                               story=st,
                               feedback = fb,
                               user_id=current_user.id, 
                               date = str(datetime.datetime.now()).split('.')[0])
            db.session.add(new_ip)
            db.session.commit()
            flash('New WebNovel added!', category='success')
            return redirect('/WebNovel')
    return render_template("adding_ip.html", user=current_user, ips=ips, max_num = max_num, all_genre = all_genre, all_ref_path = all_ref_path, all_copyright_status = all_copyright_status, this_year = this_year)

@views.route('/adding/Books', methods=['GET','POST'])
@login_required
def adding_books():
    this_year = datetime.datetime.now().year
    ips = Books.query.order_by(desc(Books.num)).all()
    max_num = get_maxnum(ips)
    all_category = db.session.query(Books.category).distinct()
    all_nation = db.session.query(Books.nation).distinct()
    all_genre = db.session.query(Books.genre).distinct()
    all_ref_path = db.session.query(Books.ref_path).distinct()
    all_copyright_status = db.session.query(Books.copyright_status).distinct()
    if request.method == 'POST':
        ip_exist = False
        n = request.form.get('num')
        t = request.form.get('title')
        cat = request.form.get('category')
        cat_added = request.form.get('category_added')
        if not cat and cat_added:
            cat = cat_added
        elif cat and cat_added:
            flash('분류를 중복으로 선택/작성 했습니다! 해당 IP에서 확인해주세요!', category='error')
        nat = request.form.get('nation')
        nat_added = request.form.get('nation_added')
        if not nat and nat_added:
            nat = nat_added
        elif nat and nat_added:
            flash('국가를 중복으로 선택/작성 했습니다! 해당 IP에서 확인해주세요!', category='error')
        g = request.form.get('genre')
        g_added = request.form.get('genre_added')
        if not g and g_added:
            g = g_added
        elif g and g_added:
            flash('주장르를 중복으로 선택/작성 했습니다! 해당 IP에서 확인해주세요!', category='error')
        kw = request.form.get('keyword')
        rf = request.form.get('ref_path')
        rf_added = request.form.get('ref_path_added')
        if not rf and rf_added:
            rf = rf_added
        elif rf and rf_added:
            flash('추천경로를 중복으로 선택/작성 했습니다! 해당 IP에서 확인해주세요!', category='error')
        c = request.form.get('copyright')
        w = request.form.get('writer')
        cs = request.form.get('copyright_status')
        cs_added = request.form.get('copyright_status_added')
        if not cs and cs_added:
            cs = cs_added
        elif cs and cs_added:
            flash('저작권/판권현황를 중복으로 선택/작성 했습니다! 해당 IP에서 확인해주세요!', category='error')
        condi = request.form.get('condition')
        dp = str(request.form.get('dp_year')) + '.' + str(request.form.get('dp_month'))
        m = request.form.get('monitored')
        dm = str(request.form.get('dm_year')) + '.' + str(request.form.get('dm_month'))
        sg = request.form.get('suggested')
        dsg = str(request.form.get('dsg_year')) + '.' + str(request.form.get('dsg_month'))
        sf = request.form.get('suggested_feedback')
        st = request.form.get('story')
        fb = request.form.get('feedback')

        if len(t) < 1:
            flash('제목이 너무 짧습니다! 확인해주세요!', category='error')
        elif len(w) < 1:
            flash('원작자가 너무 짧습니다! 확인해주세요!', category='error')
        if Books.query.filter_by(title=t).first():
            title_db = Books.query.filter_by(title=t).all()
            for ttl in title_db:
                if ttl.writer == w:
                    ip_exist = True
                    flash(f'{ttl.title} by {w} 동일한 작가의 작품이 있습니다! 확인해주세요!', category='error')
        if not ip_exist:
            new_ip = Books(num=n,
                               title = t,
                               category = cat,
                               nation = nat,
                               genre = g,
                               keyword = kw,
                               ref_path=rf,
                               copyright = c,
                               writer = w,
                               copyright_status = cs,
                               condition = condi,
                               date_pub=dp,
                               monitored = m,
                               date_monitor = dm,
                               suggested = sg,
                               date_suggested = dsg,
                               suggested_feedback = sf,
                               story=st,
                               feedback = fb,
                               user_id=current_user.id, 
                               date = str(datetime.datetime.now()).split('.')[0]
                               )
            db.session.add(new_ip)
            db.session.commit()
            flash('New Books added!', category='success')
            return redirect('/Books')
    return render_template("adding_ip.html", user=current_user, ips=ips, max_num= max_num, all_genre = all_genre, all_ref_path = all_ref_path, all_copyright_status = all_copyright_status, all_category = all_category, all_nation = all_nation, this_year = this_year)

@views.route('/delete/WebToonIP/<int:id>', methods=['GET','POST'])
@login_required
def delete_webtoon(id):
    ip_to_delete = WebToonIP.query.get_or_404(id)
    db.session.delete(ip_to_delete)
    db.session.commit()
    print ('WebToonIP', id ,'was deleted')
    return redirect('/WebToonIP')

@views.route('/delete/WebNovel/<int:id>', methods=['GET','POST'])
@login_required
def delete_webnovel(id):
    ip_to_delete = WebNovel.query.get_or_404(id)
    db.session.delete(ip_to_delete)
    db.session.commit()
    print ('WebNovel', id ,'was deleted')
    return redirect('/WebNovel')

@views.route('/delete/Books/<int:id>', methods=['GET','POST'])
@login_required
def delete_books(id):
    ip_to_delete = Books.query.get_or_404(id)
    db.session.delete(ip_to_delete)
    db.session.commit()
    print ('Books', id ,'was deleted')
    return redirect('/Books')

    
'''
@views.route('/search/WebNovel', methods=['GET', 'POST'])
@login_required
def search_webnovel():
    total_num =  len(WebNovel.query.all())
    all_genre = db.session.query(WebNovel.genre).distinct()
    all_copyright_status = db.session.query(WebNovel.copyright_status).distinct()
    all_keyword = db.session.query(WebNovel.keyword).distinct()
    this_year = datetime.datetime.now().year
    if request.method == 'POST':
        search_title = "%{}%".format(request.form.get('searchtitle'))
        dm_year = request.form.get('search_dm_year')
        dm_month = request.form.get('search_dm_month')
        if not dm_year and not dm_month:
            search_date_monitor = '%%'
        else:
            if dm_month:
                dm_month = '.' + dm_month
            search_date_monitor = '%' + dm_year + dm_month +'%'
        
        dp_year = request.form.get('search_dp_year')
        dp_month = request.form.get('search_dp_month')
        if not dp_year and not dp_month:
            search_date_pub = '%%'
        else:
            if dp_month:
                dp_month = '.' + dp_month
            search_date_pub = '%' + dp_year + dp_month +'%'
        #search_date_pub = "%{}%".format(request.form.get('search_datepub'))
        search_genre = "%{}%".format(request.form.get('search_genre'))
        search_copyright_status = "%{}%".format(request.form.get('search_copyright_status'))
        search_keyword = "%{}%".format(request.form.get('search_keyword'))
        if search_title == '%%':
            search_title = ''
        if search_date_monitor == '%%':
            search_date_monitor = Developer
        if search_date_pub == '%%':
            search_date_pub = Developer
        if search_genre == '%%':
            search_genre = Developer
        if search_copyright_status == '%%':
            search_copyright_status = Developer
        if search_keyword == '%%':
            search_keyword = Developer
        results = WebNovel.query.filter(or_(WebNovel.title.like(search_title),
                                            WebNovel.date_monitor.like(search_date_monitor),
                                            WebNovel.date_pub.like(search_date_pub),
                                            WebNovel.genre.like(search_genre),
                                            WebNovel.copyright_status.like(search_copyright_status),
                                            WebNovel.keyword.like(search_keyword)
                                            )
                                        ).all()
        total_sum = len(results)
        if total_sum != 0:
            return render_template("webtoon_card.html", user=current_user, ips = results, total_sum = total_sum, total_num = total_num, all_genre=all_genre, searched=True, all_copyright_status = all_copyright_status, all_keyword = all_keyword, this_year = this_year)
        else:
            flash('검색결과가 없습니다. 검색어를 확인해주세요!', category='error')
            return redirect('/WebNovel')
    else:
        return redirect('/WebNovel')

@views.route('/search/Books', methods=['GET', 'POST'])
@login_required
def search_books():
    total_num =  len(Books.query.all())
    all_genre = db.session.query(Books.genre).distinct()
    all_copyright_status = db.session.query(Books.copyright_status).distinct()
    all_keyword = db.session.query(Books.keyword).distinct()
    this_year = datetime.datetime.now().year
    if request.method == 'POST':
        search_title = "%{}%".format(request.form.get('searchtitle'))
        dm_year = request.form.get('search_dm_year')
        dm_month = request.form.get('search_dm_month')
        if not dm_year and not dm_month:
            search_date_monitor = '%%'
        else:
            if dm_month:
                dm_month = '.' + dm_month
            search_date_monitor = '%' + dm_year + dm_month +'%'
        
        dp_year = request.form.get('search_dp_year')
        dp_month = request.form.get('search_dp_month')
        if not dp_year and not dp_month:
            search_date_pub = '%%'
        else:
            if dp_month:
                dp_month = '.' + dp_month
            search_date_pub = '%' + dp_year + dp_month +'%'
        #search_date_pub = "%{}%".format(request.form.get('search_datepub'))
        search_genre = "%{}%".format(request.form.get('search_genre'))
        search_copyright_status = "%{}%".format(request.form.get('search_copyright_status'))
        search_keyword = "%{}%".format(request.form.get('search_keyword'))
        if search_title == '%%':
            search_title = ''
        if search_date_monitor == '%%':
            search_date_monitor = Developer
        if search_date_pub == '%%':
            search_date_pub = Developer
        if search_genre == '%%':
            search_genre = Developer
        if search_copyright_status == '%%':
            search_copyright_status = ''
        if search_keyword == '%%':
            search_keyword = Developer
        results = Books.query.filter(or_(Books.title.like(search_title),
                                         Books.date_monitor.like(search_date_monitor),
                                         Books.date_pub.like(search_date_pub),
                                         Books.genre.like(search_genre),
                                         Books.copyright_status.like(search_copyright_status),
                                         Books.keyword.like(search_keyword)
                                        )
                                        ).all()
        total_sum = len(results)
        if total_sum != 0:
            return render_template("webtoon_card.html", user=current_user, ips = results, total_sum = total_sum, total_num = total_num, all_genre=all_genre, searched=True, all_copyright_status = all_copyright_status, all_keyword = all_keyword, this_year=this_year)
        else:
            flash('검색결과가 없습니다. 검색어를 확인해주세요!', category='error')
            return redirect('/Books')
    else:
        return redirect('/Books')
'''
@views.route('/temp/WebToonIP', methods=['GET', 'POST'])
@login_required
def temp_webtoon():
    all_webtoon = WebToonIP.query.all()
    for webtoon in all_webtoon:
        if webtoon.suggested == 'N' and '2021.' in webtoon.date_suggested and len(webtoon.date_suggested) > 3:
            webtoon.suggested = 'Y'
            db.session.commit()
            print (webtoon.num, webtoon.title, webtoon.suggested, webtoon.date_suggested)
    return render_template('temp.html', user = current_user, all_ips = all_webtoon)

@views.route('/temp/WebNovel', methods=['GET', 'POST'])
@login_required
def temp_webnovel():
    all_webnovel = WebNovel.query.all()
    for webnovel in all_webnovel:
        if webnovel.suggested == 'N' and '2021.' in webnovel.date_suggested and len(webnovel.date_suggested) > 3:
            webnovel.suggested = 'Y'
            db.session.commit()
            print (webnovel.num, webnovel.title, webnovel.suggested, webnovel.date_suggested)
    return render_template('temp.html', user = current_user, all_ips = all_webnovel)

@views.route('/temp/Books', methods=['GET', 'POST'])
@login_required
def temp_books():
    all_books = Books.query.all()
    for book in all_books:
        if book.suggested == 'N' and '2021.' in book.date_suggested and len(book.date_suggested) > 3:
            book.suggested = 'Y'
            db.session.commit()
            print (book.num, book.title, book.suggested, book.date_suggested)
    return render_template('temp.html', user = current_user, all_ips = all_books)
    
@views.route('/update/WebToonIP/<int:id>', methods=['GET', 'POST'])
@login_required
def update_webtoon(id):
    this_year = datetime.datetime.now().year
    ip_to_update = WebToonIP.query.get_or_404(id)
    all_genre = db.session.query(WebToonIP.genre).distinct()
    all_ref_path = db.session.query(WebToonIP.ref_path).distinct()
    all_copyright_status = db.session.query(WebToonIP.copyright_status).distinct()
    if request.method == 'POST':
        ip_to_update.num = request.form.get('num')
        ip_to_update.title = request.form.get('title')
        g = request.form.get('genre')
        g_added = request.form.get('genre_added')
        if not g and g_added:
            g = g_added
        elif g and g_added:
            flash('주장르를 중복으로 선택/작성 했습니다! 해당 IP에서 확인해주세요!', category='error')
        ip_to_update.genre = g
        ip_to_update.keyword = request.form.get('keyword')
        rf = request.form.get('ref_path')
        rf_added = request.form.get('ref_path_added')
        if not rf and rf_added:
            rf = rf_added
        elif rf and rf_added:
            flash('추천경로를 중복으로 선택/작성 했습니다! 해당 IP에서 확인해주세요!', category='error')
        ip_to_update.ref_path = rf
        ip_to_update.copyright = request.form.get('copyright')
        ip_to_update.writer = request.form.get('writer')
        cs = request.form.get('copyright_status')
        cs_added = request.form.get('copyright_status_added')
        if not cs and cs_added:
            cs = cs_added
        elif cs and cs_added:
            flash('저작권/판권현황를 중복으로 선택/작성 했습니다! 해당 IP에서 확인해주세요!', category='error')
        ip_to_update.copyright_status = cs
        ip_to_update.condition = request.form.get('condition')
        ip_to_update.date_pub = str(request.form.get('dp_year')) + '.' + str(request.form.get('dp_month'))
        ip_to_update.monitored = request.form.get('monitored')
        ip_to_update.date_monitor = str(request.form.get('dm_year')) + '.' + str(request.form.get('dm_month'))
        ip_to_update.suggested = request.form.get('suggested')
        ip_to_update.date_suggested = str(request.form.get('dsg_year')) + '.' + str(request.form.get('dsg_month'))
        ip_to_update.story = request.form.get('story')
        ip_to_update.feedback = request.form.get('feedback')
        ip_to_update.date = str(datetime.datetime.now()).split('.')[0]
        if len(ip_to_update.title) < 1:
            flash('제목이 너무 짧습니다! 확인해주세요!', category='error')
        elif len(ip_to_update.writer) < 1:
            flash('원작자가 너무 짧습니다! 확인해주세요!', category='error')
        else:
            try:
                db.session.commit()
                flash( f'{ip_to_update.__class__.__name__} : {ip_to_update.title} Updated!', category='success')
                return render_template("detail_view.html", user=current_user, ip = ip_to_update)
            except:
                return "There was a problem updating.."
    return render_template("update.html", user=current_user, ip_to_update = ip_to_update, all_genre = all_genre, all_ref_path = all_ref_path, all_copyright_status = all_copyright_status, this_year=this_year)

@views.route('/update/WebNovel/<int:id>', methods=['GET', 'POST'])
@login_required
def update_webnovel(id):
    this_year = datetime.datetime.now().year
    all_genre = db.session.query(WebNovel.genre).distinct()
    all_ref_path = db.session.query(WebNovel.ref_path).distinct()
    all_copyright_status = db.session.query(WebNovel.copyright_status).distinct()
    ip_to_update = WebNovel.query.get_or_404(id)
    if request.method == 'POST':
        ip_to_update.num = request.form.get('num')
        ip_to_update.title = request.form.get('title')
        g = request.form.get('genre')
        g_added = request.form.get('genre_added')
        if not g and g_added:
            g = g_added
        elif g and g_added:
            flash('주장르를 중복으로 선택/작성 했습니다! 해당 IP에서 확인해주세요!', category='error')
        ip_to_update.genre = g
        ip_to_update.keyword = request.form.get('keyword')
        
        rf = request.form.get('ref_path')
        rf_added = request.form.get('ref_path_added')
        if not rf and rf_added:
            rf = rf_added
        elif rf and rf_added:
            flash('추천경로를 중복으로 선택/작성 했습니다! 해당 IP에서 확인해주세요!', category='error')
        ip_to_update.ref_path = rf
        ip_to_update.copyright = request.form.get('copyright')
        ip_to_update.writer = request.form.get('writer')
        cs = request.form.get('copyright_status')
        cs_added = request.form.get('copyright_status_added')
        if not cs and cs_added:
            cs = cs_added
        elif cs and cs_added:
            flash('저작권/판권현황를 중복으로 선택/작성 했습니다! 해당 IP에서 확인해주세요!', category='error')
        ip_to_update.copyright_status = cs
        ip_to_update.condition = request.form.get('condition')
        ip_to_update.date_pub = str(request.form.get('dp_year')) + '.' + str(request.form.get('dp_month'))
        ip_to_update.monitored = request.form.get('monitored')
        ip_to_update.date_monitor = str(request.form.get('dm_year')) + '.' + str(request.form.get('dm_month'))
        ip_to_update.suggested = request.form.get('suggested')
        ip_to_update.date_suggested = str(request.form.get('dsg_year')) + '.' + str(request.form.get('dsg_month'))
        ip_to_update.suggested_feedback = request.form.get('suggested_feedback')
        ip_to_update.story = request.form.get('story')
        ip_to_update.feedback = request.form.get('feedback')
        ip_to_update.date = str(datetime.datetime.now()).split('.')[0]
        if len(ip_to_update.title) < 1:
            flash('제목이 너무 짧습니다! 확인해주세요!', category='error')
        elif len(ip_to_update.writer) < 1:
            flash('원작자가 너무 짧습니다! 확인해주세요!', category='error')
        else:
            try:
                db.session.commit()
                flash( f'{ip_to_update.__class__.__name__} : {ip_to_update.title} Updated!', category='success')
                return render_template("detail_view.html", user=current_user, ip = ip_to_update)
            except:
                return "There was a problem updating.."
    return render_template("update.html", user=current_user, ip_to_update = ip_to_update, all_genre = all_genre, all_ref_path = all_ref_path, all_copyright_status = all_copyright_status, this_year=this_year)

@views.route('/update/Books/<int:id>', methods=['GET', 'POST'])
@login_required
def update_books(id):
    this_year = datetime.datetime.now().year
    all_category = db.session.query(Books.category).distinct()
    all_nation = db.session.query(Books.nation).distinct()
    all_genre = db.session.query(Books.genre).distinct()
    all_ref_path = db.session.query(Books.ref_path).distinct()
    all_copyright_status = db.session.query(Books.copyright_status).distinct()
    ip_to_update = Books.query.get_or_404(id)
    if request.method == 'POST':
        ip_to_update.num = request.form.get('num')
        ip_to_update.title = request.form.get('title')
        cat = request.form.get('category')
        cat_added = request.form.get('category_added')
        if not cat and cat_added:
            cat = cat_added
        elif cat and cat_added:
            flash('분류를 중복으로 선택/작성 했습니다! 해당 IP에서 확인해주세요!', category='error')
        ip_to_update.category = cat
        nat = request.form.get('nation')
        nat_added = request.form.get('nation_added')
        if not nat and nat_added:
            nat = nat_added
        elif nat and nat_added:
            flash('국가를 중복으로 선택/작성 했습니다! 해당 IP에서 확인해주세요!', category='error')
        ip_to_update.nation = nat
        g = request.form.get('genre')
        g_added = request.form.get('genre_added')
        if not g and g_added:
            g = g_added
        elif g and g_added:
            flash('주장르를 중복으로 선택/작성 했습니다! 해당 IP에서 확인해주세요!', category='error')
        ip_to_update.genre = g
        ip_to_update.keyword = request.form.get('keyword')

        rf = request.form.get('ref_path')
        rf_added = request.form.get('ref_path_added')
        if not rf and rf_added:
            rf = rf_added
        elif rf and rf_added:
            flash('추천경로를 중복으로 선택/작성 했습니다! 해당 IP에서 확인해주세요!', category='error')
        ip_to_update.ref_path = rf
        ip_to_update.copyright = request.form.get('copyright')
        ip_to_update.writer = request.form.get('writer')
        cs = request.form.get('copyright_status')
        cs_added = request.form.get('copyright_status_added')
        if not cs and cs_added:
            cs = cs_added
        elif cs and cs_added:
            flash('저작권/판권현황를 중복으로 선택/작성 했습니다! 해당 IP에서 확인해주세요!', category='error')
        ip_to_update.copyright_status = cs
        ip_to_update.condition = request.form.get('condition')
        ip_to_update.date_pub = str(request.form.get('dp_year')) + '.' + str(request.form.get('dp_month'))
        ip_to_update.monitored = request.form.get('monitored')
        ip_to_update.date_monitor = str(request.form.get('dm_year')) + '.' + str(request.form.get('dm_month'))
        ip_to_update.suggested = request.form.get('suggested')
        ip_to_update.date_suggested = str(request.form.get('dsg_year')) + '.' + str(request.form.get('dsg_month'))
        ip_to_update.suggested_feedback = request.form.get('suggested_feedback')
        ip_to_update.story = request.form.get('story')
        ip_to_update.feedback = request.form.get('feedback')
        ip_to_update.date = str(datetime.datetime.now()).split('.')[0]
        if len(ip_to_update.title) < 1:
            flash('제목이 너무 짧습니다! 확인해주세요!', category='error')
        elif len(ip_to_update.writer) < 1:
            flash('원작자가 너무 짧습니다! 확인해주세요!', category='error')
        else:
            try:
                db.session.commit()
                flash( f'{ip_to_update.__class__.__name__} : {ip_to_update.title} Updated!', category='success')
                return render_template("detail_view.html", user=current_user, ip = ip_to_update)
            except:
                return "There was a problem updating.."
    return render_template("update.html", user=current_user, ip_to_update = ip_to_update, all_genre = all_genre, all_ref_path = all_ref_path, all_copyright_status = all_copyright_status, all_category = all_category, all_nation = all_nation, this_year = this_year)

@views.route('/delete-note', methods=['POST'])
def delete_note():
    note = json.loads(request.data)
    noteId = note['noteId']
    note = WebToonIP.query.get(noteId)
    if note:
        #if note.user_id == current_user.id:
        db.session.delete(note)
        db.session.commit()
    return jsonify({})
